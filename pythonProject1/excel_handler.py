from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from measurements import Measurements


def create_file(file_path):
    wb = Workbook()
    ws = wb.active
    titles = ["Number", "Date", "Time", "category", "Speed", "Tair", "Troad", "Wind speed", "La,max", "63",
              "80", "100", "125", "160", "200", "250", "315", "400", "500", "630", "800", "1000", "1250", "1600",
              "2000", "2500", "3150", "4000", "5000", "6300", "8000", "10000", "12500", "16000", "20000"]
    ws.append(titles)
    wb.save(file_path)
    print("file saved succesfully")


def save_measurement(file_path, measurements):
    wb = load_workbook(file_path)
    ws = wb.active

    for m in measurements:
        last_empty_row = len(list(ws.rows))
        row = [last_empty_row, m.date, m.time, m.category, m.speed, m.air_temp, m.road_temp, m.wind_speed, m.la_max,
               m.octave[10], m.octave[11], m.octave[12], m.octave[13], m.octave[14], m.octave[15], m.octave[16],
               m.octave[17], m.octave[18], m.octave[19], m.octave[20], m.octave[21], m.octave[22], m.octave[23],
               m.octave[24], m.octave[25], m.octave[26], m.octave[27], m.octave[28], m.octave[29], m.octave[30],
               m.octave[31], m.octave[32], m.octave[33], m.octave[34], m.octave[35]]
        ws.append(row)

    wb.save(file_path)
